"""
优点:
1. 操作方面
2. 功能更强大

缺点：
不支持xls
"""
import datetime
import random
import openpyxl  # pip install openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

# 加载
wb = openpyxl.load_workbook('day025/data.xlsx')

# 获取sheet名称
print(wb.sheetnames)

# 获取sheet
sheet = wb.worksheets[0]

# 单元格数据范围
print(sheet.dimensions)

# 读取单元格数据
for row_ch in range(2, sheet.max_row + 1):
    for col_ch in 'ABCD':
        value = sheet[f'{col_ch}{row_ch}'].value
        if type(value) == datetime.datetime:
            print(value.strftime('%Y年%m月%d日'), end='\t')
        elif type(value) == int:
            print(f'{value:<10d}', end='\t')  # <10d, 从左开始，展示10的宽度的，10进制整数
        elif type(value) == float:
            print(f'{value:.4f}', end='\t')
        else:
            print(value, end='\t')
    print()


# 写
wb2 = openpyxl.Workbook()
sheet = wb.active
sheet.title = '期末考试'
titles = ('姓名', '语文', '数学', '英语')
for col_index, title in enumerate(titles):
    sheet.cell(1, col_index + 1, title)  # cell(row, col , value)

names = ('关羽', '张飞', '赵云', '马超', '黄忠')
for row_index, name in enumerate(names):
    sheet.cell(row_index + 2, 1, name)
    for col_index in range(2, 5):
        sheet.cell(row_index + 2, col_index, random.randrange(50, 101))

wb.save('day025/考试成绩表.xlsx')


# 格式
alignment = Alignment(horizontal='center', vertical='center')   # 对齐方式
side = Side(color='ff7f50', style='mediumDashed')               # 边框线条
border = Border(left=side, top=side, right=side, bottom=side)   # 边框
font = Font(size=18, bold=True, color='ff1493', name='华文楷体')    # 字体

wb2 = openpyxl.Workbook()
sheet = wb.active

sheet.cell(1, 5).font = font
sheet.cell(1, 5).alignment = alignment
sheet.cell(1, 5).border = border

sheet.row_dimensions[1].height = 30  # 设置行高
sheet.column_dimensions['E'].width = 120  # 设置列宽

sheet['E1'] = '平均分'


sheet.title = '期末考试'
titles = ('姓名', '语文', '数学', '英语')
for col_index, title in enumerate(titles):
    sheet.cell(1, col_index + 1, title)  # cell(row, col , value)

names = ('关羽', '张飞', '赵云', '马超', '黄忠')
for row_index, name in enumerate(names):
    sheet.cell(row_index + 2, 1, name)
    for col_index in range(2, 5):
        sheet.cell(row_index + 2, col_index, random.randrange(50, 101))

for i in range(2, 7):
    sheet[f'E{i}'] = f'=average(B{i}:D{i})'
    sheet.cell(i, 5).font = Font(size=12, color='4169e1', italic=True)
    sheet.cell(i, 5).alignment = alignment

wb.save('day025/考试成绩表（格式）.xlsx')

# 统计表
wb3 = openpyxl.Workbook(write_only=True)
sheet = wb3.create_sheet()

rows = [
    ('类别', '销售A组', '销售B组'),
    ('手机', 40, 30),
    ('平板', 50, 60),
    ('笔记本', 80, 70),
    ('外围设备', 20, 10),
]

for row in rows:
    sheet.append(row)  # append会另起一行追加

# 图表
chart = BarChart()  # 柱状图
chart.type = 'col'  # 垂直方向
chart.style = 10    # 预定义的样式编号
chart.shape = 4     # 柱子的形状样式

chart.title = '销售统计图'
chart.y_axis.title = '销量'
chart.x_axis.title = '商品类别'


# 数据范围
data = Reference(sheet, min_row=1, max_row=5, min_col=2, max_col=3)

# 分类范围
cats = Reference(sheet,  min_row=2, max_row=5, min_col=1)  # 不设置max_col，就代表是1列

# 给图表添加数据
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
sheet.add_chart(chart, 'A10')
wb3.save('day025/图表演示.xlsx')
